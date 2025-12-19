"""
Excel 导入服务模块
处理从 Excel 文件导入 Creator 数据到 Supabase 的业务逻辑
"""
import os
import json
import hashlib
import logging
import requests
import uuid
import boto3
from typing import List, Dict, Any, Tuple, Optional
from io import BytesIO
from datetime import datetime
from PIL import Image
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed

logger = logging.getLogger(__name__)


def load_env_file(path: str = ".env"):
    """Load environment variables from a .env file (supports `export KEY=value`).
    
    First tries the provided path, then tries relative to the script directory,
    then tries the current working directory.
    """
    # 获取脚本所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)  # 项目根目录
    
    # 尝试多个可能的路径
    possible_paths = [
        path,  # 用户提供的路径
        os.path.join(script_dir, path),  # 相对于脚本目录
        os.path.join(project_root, path),  # 相对于项目根目录
        os.path.join(os.getcwd(), path),  # 相对于当前工作目录
    ]
    
    env_file = None
    for p in possible_paths:
        if os.path.exists(p):
            env_file = p
            break
    
    if not env_file:
        logger.debug(f"未找到 .env 文件（已检查: {', '.join(possible_paths)}）")
        return
    
    logger.debug(f"加载 .env 文件: {env_file}")
    with open(env_file, "r", encoding="utf-8") as file:
        for raw_line in file:
            line = raw_line.strip()
            if not line or line.startswith("#"):
                continue
            if line.startswith("export "):
                line = line[len("export ") :]
            if "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"\'')
            os.environ[key] = value


# 在模块加载时加载 .env 文件
load_env_file()


class ExcelImportError(Exception):
    """Excel 导入相关异常"""
    pass


class DifyAPIError(Exception):
    """Dify API 调用失败异常"""
    pass


class ExcelImporter:
    """Excel 导入服务类"""
    
    def __init__(self, supabase_url: str, supabase_api_key: str, dify_url: str = None, dify_api_key: str = None, dify_user: str = "excel-importer", s3_bucket: str = None, s3_key_prefix: str = None):
        """
        初始化 Excel 导入器
        
        Args:
            supabase_url: Supabase URL
            supabase_api_key: Supabase API Key
            dify_url: Dify API URL (可选，从环境变量 DIFY_URL 获取)
            dify_api_key: Dify API Key (可选，从环境变量 DIFY_API_KEY_TOKEN 获取)
            dify_user: Dify API User (可选，从环境变量 DIFY_USER 获取，默认 "excel-importer")
        """
        self.supabase_url = supabase_url
        self.supabase_api_key = supabase_api_key
        self.api_url = f"{supabase_url.rstrip('/')}/rest/v1/creator"
        self.headers = {
            "apikey": supabase_api_key,
            "Authorization": f"Bearer {supabase_api_key}",
            "Content-Type": "application/json",
            "Prefer": "return=representation"
        }
        
        # Dify 配置
        # 必须从环境变量 DIFY_API_KEY_TOKEN 获取
        self.dify_url = dify_url or os.getenv("DIFY_URL")
        self.dify_api_key = dify_api_key or os.getenv("DIFY_API_KEY_TOKEN")
        self.dify_user = dify_user or os.getenv("DIFY_USER", "excel-importer")
        
        # S3 配置
        self.s3_bucket = s3_bucket or os.getenv("S3_BUCKET") or os.getenv("S3_BUCKET_NAME") or "amzn-s3-fc-bucket"
        self.s3_key_prefix = s3_key_prefix or os.getenv("S3_CREATOR_PREFIX") or "creator-signatures"
        
        # 记录 Dify 配置状态
        if self.dify_url and self.dify_api_key:
            logger.info(f"Dify 配置已设置: URL={self.dify_url[:50]}..., User={self.dify_user}")
        else:
            missing = []
            if not self.dify_url:
                missing.append("DIFY_URL")
            if not self.dify_api_key:
                missing.append("DIFY_API_KEY_TOKEN")
            # 调试信息：检查环境变量是否存在
            env_dify_url = os.getenv("DIFY_URL")
            env_dify_api_key_token = os.getenv("DIFY_API_KEY_TOKEN")
            logger.warning(f"Dify 配置未完整设置，将跳过 tokens 生成。缺少: {', '.join(missing)}")
            logger.debug(f"环境变量检查: DIFY_URL={'已设置' if env_dify_url else '未设置'}, "
                        f"DIFY_API_KEY_TOKEN={'已设置' if env_dify_api_key_token else '未设置'}")
        
        # 记录 S3 配置状态
        logger.info(f"S3 配置: Bucket={self.s3_bucket}, KeyPrefix={self.s3_key_prefix}")
    
    def _download_image(self, url: str) -> Image.Image:
        """
        从 URL 下载图片
        
        Args:
            url: 图片 URL
            
        Returns:
            PIL Image 对象
            
        Raises:
            ExcelImportError: 如果下载失败
        """
        try:
            logger.debug(f"    开始下载图片: {url}")
            response = requests.get(url, timeout=30, stream=True)
            response.raise_for_status()
            
            # 检查 Content-Type 是否为图片
            content_type = response.headers.get("Content-Type", "").lower()
            if not content_type.startswith("image/"):
                logger.warning(f"    警告: URL 的 Content-Type 不是图片类型: {content_type}")
            
            image = Image.open(BytesIO(response.content)).convert("RGB")
            logger.debug(f"    ✓ 图片下载成功，尺寸: {image.size}")
            return image
        except requests.exceptions.RequestException as e:
            error_msg = f"下载图片失败: {url} - {str(e)}"
            logger.error(f"    {error_msg}")
            raise ExcelImportError(error_msg) from e
        except Exception as e:
            error_msg = f"处理图片失败: {url} - {str(e)}"
            logger.error(f"    {error_msg}")
            raise ExcelImportError(error_msg) from e
    
    def _build_s3_object_key(self, prefix: str, extension: str = "png") -> str:
        """
        构建 S3 对象键
        
        Args:
            prefix: 键前缀
            extension: 文件扩展名（不含点）
            
        Returns:
            S3 对象键
        """
        clean_prefix = prefix.strip("/ ")
        date_path = datetime.utcnow().strftime("%Y/%m/%d")
        filename = f"{uuid.uuid4().hex}.{extension}"
        parts = [segment for segment in (clean_prefix, date_path, filename) if segment]
        return "/".join(parts)
    
    def _upload_image_to_s3(self, image: Image.Image, content_type: str = "image/png", extension: str = "png") -> str:
        """
        将 PIL Image 上传到 S3
        
        Args:
            image: PIL Image 对象
            content_type: 内容类型
            extension: 文件扩展名（不含点）
            
        Returns:
            S3 公网 URL
            
        Raises:
            ExcelImportError: 如果上传失败
        """
        try:
            # 将图片保存到内存缓冲区
            buffer = BytesIO()
            # PIL 只接受 "JPEG" 作为格式，不接受 "JPG"，需要映射
            format_map = {
                "PNG": "PNG",
                "JPEG": "JPEG",
                "JPG": "JPEG",  # JPG 映射到 JPEG
                "GIF": "GIF",
                "WEBP": "WEBP"
            }
            pil_format = format_map.get(extension.upper(), "PNG")
            image.save(buffer, format=pil_format)
            buffer.seek(0)
            
            # 获取 AWS 凭证和区域
            aws_access_key_id = os.environ.get("AWS_ACCESS_KEY_ID")
            aws_secret_access_key = os.environ.get("AWS_SECRET_ACCESS_KEY")
            region = (
                os.environ.get("AWS_REGION")
                or os.environ.get("AWS_DEFAULT_REGION")
                or "us-east-1"
            )
            
            # 创建 boto3 Session，如果环境变量中有凭证则显式传入
            if aws_access_key_id and aws_secret_access_key:
                session = boto3.session.Session(
                    aws_access_key_id=aws_access_key_id,
                    aws_secret_access_key=aws_secret_access_key,
                    region_name=region
                )
            else:
                # 如果没有显式凭证，尝试使用默认凭证链（环境变量、配置文件、IAM 角色等）
                session = boto3.session.Session(region_name=region)
            
            # 创建 S3 客户端
            s3_client = session.client("s3", region_name=region)
            
            # 构建对象键
            object_key = self._build_s3_object_key(self.s3_key_prefix, extension)
            
            logger.debug(f"    上传图片到 S3: Bucket={self.s3_bucket}, Key={object_key}")
            
            # 上传到 S3
            s3_client.upload_fileobj(
                buffer,
                self.s3_bucket,
                object_key,
                ExtraArgs={"ACL": "public-read", "ContentType": content_type},
            )
            
            # 构建公网 URL
            if region == "us-east-1":
                s3_url = f"https://{self.s3_bucket}.s3.amazonaws.com/{object_key}"
            else:
                s3_url = f"https://{self.s3_bucket}.s3.{region}.amazonaws.com/{object_key}"
            
            logger.debug(f"    ✓ 图片上传成功: {s3_url}")
            return s3_url
            
        except Exception as e:
            error_msg = f"上传图片到 S3 失败: {str(e)}"
            logger.error(f"    {error_msg}")
            raise ExcelImportError(error_msg) from e
    
    def _download_and_upload_image_to_s3(self, image_url: str) -> Optional[str]:
        """
        从 URL 下载图片并上传到 S3
        
        Args:
            image_url: 图片 URL
            
        Returns:
            S3 公网 URL，如果失败则返回 None（不会抛出异常）
        """
        if not image_url or not image_url.strip():
            return None
        
        try:
            # 下载图片
            image = self._download_image(image_url)
            
            # 确定文件扩展名和内容类型
            # 尝试从 URL 获取扩展名
            url_lower = image_url.lower()
            if url_lower.endswith((".jpg", ".jpeg")):
                extension = "jpg"
                content_type = "image/jpeg"
            elif url_lower.endswith(".png"):
                extension = "png"
                content_type = "image/png"
            elif url_lower.endswith(".gif"):
                extension = "gif"
                content_type = "image/gif"
            elif url_lower.endswith(".webp"):
                extension = "webp"
                content_type = "image/webp"
            else:
                # 默认使用 PNG
                extension = "png"
                content_type = "image/png"
            
            # 上传到 S3
            s3_url = self._upload_image_to_s3(image, content_type, extension)
            return s3_url
            
        except Exception as e:
            # 所有异常都捕获，记录警告但不中断导入流程
            # 这样即使图片处理失败，也不会跳过整个 Creator
            logger.warning(f"    处理图片失败（将使用原始 URL）: {image_url} - {str(e)}")
            return None
    
    def validate_file(self, filename: str, contents: bytes) -> None:
        """
        验证文件格式
        
        Args:
            filename: 文件名
            contents: 文件内容
            
        Raises:
            ExcelImportError: 如果文件格式不正确
        """
        if not filename:
            raise ExcelImportError("未提供文件名")
        
        if not filename.endswith(('.xlsx', '.xls')):
            raise ExcelImportError("文件必须是 Excel 格式 (.xlsx 或 .xls)")
        
        if not contents or len(contents) == 0:
            raise ExcelImportError("文件为空或无法读取")
    
    def parse_excel(self, contents: bytes) -> Tuple[List[str], Dict[str, int], List[List[Any]]]:
        """
        解析 Excel 文件
        
        Args:
            contents: Excel 文件内容
            
        Returns:
            (headers, column_map, rows): 表头列表、列映射字典、数据行列表
            
        Raises:
            ExcelImportError: 如果解析失败
        """
        try:
            logger.info("开始解析 Excel 文件...")
            workbook = load_workbook(filename=BytesIO(contents), data_only=True)
            logger.info(f"✓ Excel 文件解析成功，工作表数量: {len(workbook.sheetnames)}")
            logger.info(f"工作表名称: {workbook.sheetnames}")
        except Exception as e:
            logger.error(f"解析 Excel 文件失败: {str(e)}", exc_info=True)
            raise ExcelImportError(f"解析 Excel 文件失败: {str(e)}。请确保文件格式正确。") from e
        
        sheet = workbook.active
        logger.info(f"使用活动工作表: {sheet.title}")
        logger.info(f"工作表最大行数: {sheet.max_row}, 最大列数: {sheet.max_column}")
        
        # 获取表头（第一行）
        logger.info("开始解析表头...")
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value if cell.value else "")
        logger.info(f"✓ 表头解析完成，共 {len(headers)} 列")
        logger.info(f"表头内容: {headers}")
        
        # 映射列索引
        column_map = self._map_columns(headers)
        
        # 检查必需字段
        if "creator_name" not in column_map:
            raise ExcelImportError("Excel 文件中未找到 'Creator' 列")
        
        # 读取数据行
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=False):
            rows.append(row)
        
        return headers, column_map, rows
    
    def _map_columns(self, headers: List[str]) -> Dict[str, int]:
        """
        映射列名到列索引
        
        Args:
            headers: 表头列表
            
        Returns:
            列名到列索引的映射字典
        """
        logger.info("开始映射列索引...")
        column_map = {}
        
        for idx, header in enumerate(headers, start=1):
            header_str = str(header).strip() if header else ""
            header_lower = header_str.lower()
            
            # creator_id 列
            if header_str.lower() == "creator_id" or "creator_id" in header_lower:
                if "creator_id" not in column_map:
                    column_map["creator_id"] = idx
                    logger.info(f"  → 找到 'creator_id' 列，索引: {idx}")
            # creator_name 列
            elif header_str.lower() == "creator_name" or header_str == "Creator" or "creator_name" in header_lower:
                if "creator_name" not in column_map:
                    column_map["creator_name"] = idx
                    logger.info(f"  → 找到 'creator_name' 列，索引: {idx}")
            # region 列（国家）
            elif header_str.lower() == "region" or header_str == "国家" or "region" in header_lower:
                if "region" not in column_map:
                    column_map["region"] = idx
                    logger.info(f"  → 找到 'region' 列，索引: {idx}")
            # newsletter_name 列
            elif header_str.lower() == "newsletter_name" or "Newsletter 名称" in header_str or "newsletter_name" in header_lower:
                if "newsletter_name" not in column_map:
                    column_map["newsletter_name"] = idx
                    logger.info(f"  → 找到 'newsletter_name' 列，索引: {idx}")
            # contact_email 列
            elif header_str.lower() == "contact_email" or "联系邮箱" in header_str or "contact_email" in header_lower or ("邮箱" in header_str and "联系" in header_str):
                if "contact_email" not in column_map:
                    column_map["contact_email"] = idx
                    logger.info(f"  → 找到 'contact_email' 列，索引: {idx}")
            # website_url 列
            elif header_str.lower() == "website_url" or "Substack主页" in header_str or "website_url" in header_lower or "substack" in header_lower:
                if "website_url" not in column_map:
                    column_map["website_url"] = idx
                    logger.info(f"  → 找到 'website_url' 列，索引: {idx}")
            # creator_signature_image_url 列
            elif header_str.lower() == "creator_signature_image_url" or "creator_signature_image_url" in header_lower or "签名图片" in header_str or "头像" in header_str:
                if "creator_signature_image_url" not in column_map:
                    column_map["creator_signature_image_url"] = idx
                    logger.info(f"  → 找到 'creator_signature_image_url' 列，索引: {idx}")
            # platform 列
            elif header_str.lower() == "platform" or header_str == "平台" or "platform" in header_lower:
                if "platform" not in column_map:
                    column_map["platform"] = idx
                    logger.info(f"  → 找到 'platform' 列，索引: {idx}")
            # content_category 列
            elif header_str.lower() == "content_category" or "主要领域" in header_str or "content_category" in header_lower or ("领域" in header_str and "主要" in header_str):
                if "content_category" not in column_map:
                    column_map["content_category"] = idx
                    logger.info(f"  → 找到 'content_category' 列，索引: {idx}")
        
        logger.info(f"✓ 列映射完成，映射结果: {column_map}")
        return column_map
    
    def parse_data_rows(self, rows: List[List[Any]], column_map: Dict[str, int]) -> List[Dict[str, Any]]:
        """
        解析数据行，转换为 Creator 数据字典列表
        
        Args:
            rows: Excel 数据行列表
            column_map: 列映射字典
            
        Returns:
            Creator 数据字典列表
        """
        logger.info("开始解析数据行...")
        creators_to_insert = []
        skipped_rows = []
        processed_rows = 0
        
        for row_idx, row in enumerate(rows, start=2):
            processed_rows += 1
            if processed_rows % 10 == 0:
                logger.info(f"  处理进度: 已处理 {processed_rows} 行，已解析 {len(creators_to_insert)} 条有效数据")
            
            # 跳过空行
            creator_name = row[column_map["creator_name"] - 1].value if column_map["creator_name"] <= len(row) else None
            if not creator_name or str(creator_name).strip() == "":
                logger.debug(f"  第 {row_idx} 行: Creator 名称为空，跳过")
                skipped_rows.append(row_idx)
                continue
            
            # 构建 creator 数据
            try:
                creator_data = self._build_creator_data(row, row_idx, column_map)
                creators_to_insert.append(creator_data)
                logger.debug(f"    ✓ 第 {row_idx} 行数据解析完成")
            except Exception as e:
                # 其他异常也跳过
                creator_name = row[column_map["creator_name"] - 1].value if column_map["creator_name"] <= len(row) else "Unknown"
                logger.warning(f"  第 {row_idx} 行: Creator '{creator_name}' 处理失败，跳过: {str(e)}")
                skipped_rows.append(row_idx)
                continue
        
        logger.info(f"✓ 数据行解析完成")
        logger.info(f"  总处理行数: {processed_rows}")
        logger.info(f"  有效数据行数: {len(creators_to_insert)}")
        logger.info(f"  跳过行数: {len(skipped_rows)}")
        if skipped_rows:
            logger.info(f"  跳过的行号: {skipped_rows[:10]}{'...' if len(skipped_rows) > 10 else ''}")
        
        if not creators_to_insert:
            raise ExcelImportError("Excel 文件中没有有效的数据行")
        
        # 并发生成 tokens（一次并发10个请求）
        logger.info("=" * 80)
        logger.info("开始并发生成 tokens...")
        creators_with_tokens = self._generate_tokens_batch_concurrent(creators_to_insert, max_workers=10)
        logger.info("=" * 80)
        
        return creators_with_tokens
    
    def _build_creator_data(self, row: List[Any], row_idx: int, column_map: Dict[str, int]) -> Dict[str, Any]:
        """
        从 Excel 行构建 Creator 数据字典
        
        Args:
            row: Excel 行数据
            row_idx: 行索引（用于日志）
            column_map: 列映射字典
            
        Returns:
            Creator 数据字典
        """
        creator_name = row[column_map["creator_name"] - 1].value
        creator_name_str = str(creator_name).strip()
        creator_data = {
            "creator_name": creator_name_str,
        }
        
        # 尝试从 creator_id 列获取 creator_id
        if "creator_id" in column_map and column_map["creator_id"] <= len(row):
            creator_id_value = row[column_map["creator_id"] - 1].value
            if creator_id_value is not None:
                creator_id = str(creator_id_value).strip()
                if creator_id:
                    creator_data["creator_id"] = creator_id
                    logger.debug(f"    → 使用 creator_id 列: {creator_id}")
        
        # 如果没有从 creator_id 列获取到，生成一个基于 creator_name 和 email 的稳定 ID
        if "creator_id" not in creator_data:
            # 使用 creator_name 和 contact_email（如果有）生成一个稳定的 ID
            id_source = creator_name_str.lower().strip()
            if "contact_email" in column_map and column_map["contact_email"] <= len(row):
                email = row[column_map["contact_email"] - 1].value
                if email:
                    id_source += "|" + str(email).strip().lower()
            
            # 生成一个基于内容的稳定哈希值（前16位）
            creator_id_hash = hashlib.md5(id_source.encode('utf-8')).hexdigest()[:16]
            creator_data["creator_id"] = creator_id_hash
            logger.debug(f"    → 生成基于内容的 creator_id: {creator_id_hash} (来源: {id_source[:50]}...)")
        
        logger.debug(f"  第 {row_idx} 行: 处理 Creator '{creator_name_str}' (creator_id: {creator_data.get('creator_id')})")
        
        # 添加可选字段
        if "region" in column_map and column_map["region"] <= len(row):
            region = row[column_map["region"] - 1].value
            if region:
                creator_data["region"] = str(region).strip()
                logger.debug(f"    → 地区: {creator_data['region']}")
        
        if "newsletter_name" in column_map and column_map["newsletter_name"] <= len(row):
            newsletter_name = row[column_map["newsletter_name"] - 1].value
            if newsletter_name:
                creator_data["newsletter_name"] = str(newsletter_name).strip()
                logger.debug(f"    → Newsletter: {creator_data['newsletter_name']}")
        
        if "contact_email" in column_map and column_map["contact_email"] <= len(row):
            contact_email = row[column_map["contact_email"] - 1].value
            if contact_email:
                creator_data["contact_email"] = str(contact_email).strip()
                logger.debug(f"    → 邮箱: {creator_data['contact_email']}")
        
        if "website_url" in column_map and column_map["website_url"] <= len(row):
            website_url = row[column_map["website_url"] - 1].value
            if website_url:
                creator_data["website_url"] = str(website_url).strip()
                logger.debug(f"    → 网站: {creator_data['website_url']}")
                # 如果 website_url 是 Substack 主页，也设置 paid_subscribe_url
                if "substack" in str(website_url).lower():
                    creator_data["paid_subscribe_url"] = str(website_url).strip()
                    logger.debug(f"    → 检测到 Substack 链接，设置 paid_subscribe_url")
        
        if "platform" in column_map and column_map["platform"] <= len(row):
            platform = row[column_map["platform"] - 1].value
            if platform:
                creator_data["platform"] = str(platform).strip()
                logger.debug(f"    → 平台: {creator_data['platform']}")
        
        if "content_category" in column_map and column_map["content_category"] <= len(row):
            category = row[column_map["content_category"] - 1].value
            if category:
                categories = self._parse_categories(category)
                if categories:
                    creator_data["content_category"] = categories
                    logger.debug(f"    → 分类: {categories}")
                else:
                    creator_data["content_category"] = []
            else:
                creator_data["content_category"] = []
        else:
            # 如果没有 content_category 列，设置为空列表
            creator_data["content_category"] = []
        
        if "creator_signature_image_url" in column_map and column_map["creator_signature_image_url"] <= len(row):
            image_url = row[column_map["creator_signature_image_url"] - 1].value
            if image_url:
                original_url = str(image_url).strip()
                logger.debug(f"    → 原始签名图片URL: {original_url}")
                
                # 尝试下载图片并上传到 S3
                # _download_and_upload_image_to_s3 不会抛出异常，失败时返回 None
                s3_url = self._download_and_upload_image_to_s3(original_url)
                if s3_url:
                    creator_data["creator_signature_image_url"] = s3_url
                    logger.info(f"    ✓ 签名图片已上传到 S3: {s3_url}")
                else:
                    # 如果上传失败，使用原始 URL
                    creator_data["creator_signature_image_url"] = original_url
                    logger.warning(f"    ⚠ 签名图片上传到 S3 失败，使用原始 URL: {original_url}")
        
        # 注意：Dify API 调用已移至 parse_data_rows 方法中批量并发处理
        # 这里只返回基础数据，tokens 将在后续批量生成
        
        return creator_data
    
    def _parse_categories(self, category: Any) -> List[str]:
        """
        解析分类字段，支持多种分隔符
        
        Args:
            category: 分类值（可能是字符串、列表等）
            
        Returns:
            分类列表
        """
        categories = []
        
        # 如果已经是列表/数组格式
        if isinstance(category, (list, tuple)):
            for item in category:
                item_str = str(item).strip()
                if item_str:
                    # 处理 " / " 分隔符
                    if " / " in item_str:
                        split_items = [c.strip() for c in item_str.split(" / ") if c.strip()]
                        categories.extend(split_items)
                    else:
                        categories.append(item_str)
        else:
            # 字符串格式，处理分类，可能是逗号、中文顿号、分号或斜杠分隔的字符串
            category_str = str(category).strip()
            # 统一替换各种分隔符为英文逗号（注意顺序：先处理 " / "，再处理其他）
            category_str = category_str.replace(" / ", ",").replace("、", ",").replace(";", ",")
            # 分割并清理
            categories = [c.strip() for c in category_str.split(",") if c.strip()]
        
        return categories
    
    def _call_dify_for_tokens(self, creator_data: Dict[str, Any], token_type: str = "direct") -> List[str]:
        """
        调用 Dify API 生成 creator tokens
        
        Args:
            creator_data: Creator 数据字典
            token_type: token 类型，"direct" 或 "implied"
            
        Returns:
            tokens 列表
            
        Raises:
            ExcelImportError: 如果调用失败
        """
        if not self.dify_url or not self.dify_api_key:
            logger.debug(f"    Dify 配置未设置，跳过生成 {token_type} tokens")
            return []
        
        try:
            # 构建 API URL
            api_url = f"{self.dify_url.rstrip('/')}/v1/chat-messages"
            
            # 构建 query 字符串
            # 根据用户提供的 curl 示例格式构建
            creator_name = creator_data.get("creator_name", "")
            newsletter_name = creator_data.get("newsletter_name", "")
            content_category = creator_data.get("content_category", [])
            
            # 格式化 content_category 为 JSON 数组字符串
            category_str = ""
            if content_category:
                if isinstance(content_category, list):
                    # 格式化为多行 JSON 数组格式，如用户示例所示
                    category_lines = json.dumps(content_category, ensure_ascii=False, indent=2)
                    category_str = category_lines
                else:
                    category_str = str(content_category)
            
            # 构建 query（按照用户提供的格式）
            query_parts = []
            if creator_name:
                query_parts.append(f"creator_name: '{creator_name}'")
            if newsletter_name:
                query_parts.append(f"newsletter_name: '{newsletter_name}'")
            if category_str:
                # 注意：用户示例中使用了中文冒号，这里保持英文冒号
                query_parts.append(f"content_category: {category_str}")
            
            query = ",".join(query_parts) if query_parts else "creator_name: ''"
            
            # 根据 token_type 调整 query
            # 如果需要为 direct 和 implied 使用不同的 prompt，可以在这里调整
            # 目前先使用相同的 query，后续可以根据实际需求调整
            
            # 设置请求头
            headers = {
                "Authorization": f"Bearer {self.dify_api_key}",
                "Content-Type": "application/json"
            }
            
            # 构建请求体（使用 blocking 模式）
            payload = {
                "inputs": {},
                "query": query,
                "response_mode": "blocking",
                "conversation_id": "",
                "user": self.dify_user
            }
            
            logger.info(f"      调用 Dify API 获取 {token_type} tokens...")
            logger.debug(f"      Query: {query}")
            
            # 发送 POST 请求（阻塞式，不使用 stream）
            response = requests.post(api_url, headers=headers, json=payload, timeout=60)
            
            # 如果响应状态码不是 2xx，记录详细错误信息并返回
            if not response.ok:
                error_detail = f"HTTP {response.status_code} {response.reason}"
                try:
                    error_body = response.json()
                    error_detail += f"\n响应内容: {json.dumps(error_body, ensure_ascii=False, indent=2)}"
                except (json.JSONDecodeError, ValueError):
                    error_detail += f"\n响应内容: {response.text[:500]}"
                
                error_detail += f"\n请求 URL: {api_url}"
                error_detail += f"\n请求 Payload: {json.dumps(payload, ensure_ascii=False, indent=2)}"
                
                logger.error(f"    Dify API 调用失败（{token_type} tokens）: {error_detail}")
                # 抛出异常，让调用者知道失败并跳过当前 Creator
                raise DifyAPIError(f"Dify API 调用失败（{token_type} tokens）: {error_detail}")
            
            # 解析 JSON 响应
            try:
                response_data = response.json()
                
                # Dify blocking 模式返回的格式可能是：
                # 1. 直接返回数组: ["token1", "token2"]
                # 2. 包含 answer 字段: {"answer": "[\"token1\", \"token2\"]"} 或 {"answer": "{\"creator_tokens_direct\": [...], \"creator_tokens_implied\": [...]}"}
                # 3. 包含 message 字段: {"message": {"answer": "..."}}
                
                tokens = None
                
                # 情况1: 直接是数组
                if isinstance(response_data, list):
                    tokens = response_data
                # 情况2: 包含 answer 字段
                elif isinstance(response_data, dict):
                    answer = None
                    
                    # 尝试从不同位置获取 answer
                    if "answer" in response_data:
                        answer = response_data["answer"]
                    elif "message" in response_data:
                        message = response_data["message"]
                        if isinstance(message, dict) and "answer" in message:
                            answer = message["answer"]
                        elif isinstance(message, str):
                            answer = message
                    
                    if answer:
                        # answer 可能是字符串（JSON 格式）或已经是数组/对象
                        if isinstance(answer, str):
                            try:
                                # 尝试解析为 JSON
                                parsed_answer = json.loads(answer)
                                
                                # 如果解析后是对象，尝试提取对应的 tokens 字段
                                if isinstance(parsed_answer, dict):
                                    # 根据 token_type 提取对应的字段
                                    token_key = f"creator_tokens_{token_type}"
                                    if token_key in parsed_answer:
                                        tokens = parsed_answer[token_key]
                                    else:
                                        # 如果没有找到对应的字段，尝试直接查找数组字段
                                        for key, value in parsed_answer.items():
                                            if isinstance(value, list) and "token" in key.lower():
                                                tokens = value
                                                break
                                elif isinstance(parsed_answer, list):
                                    tokens = parsed_answer
                            except json.JSONDecodeError:
                                # 如果解析失败，尝试提取 JSON 数组或对象
                                # 先尝试提取数组
                                start_idx = answer.find("[")
                                end_idx = answer.rfind("]")
                                if start_idx != -1 and end_idx != -1:
                                    try:
                                        tokens = json.loads(answer[start_idx:end_idx + 1])
                                    except json.JSONDecodeError:
                                        pass
                                
                                # 如果数组提取失败，尝试提取对象
                                if not isinstance(tokens, list):
                                    start_idx = answer.find("{")
                                    end_idx = answer.rfind("}")
                                    if start_idx != -1 and end_idx != -1:
                                        try:
                                            parsed_obj = json.loads(answer[start_idx:end_idx + 1])
                                            if isinstance(parsed_obj, dict):
                                                token_key = f"creator_tokens_{token_type}"
                                                if token_key in parsed_obj:
                                                    tokens = parsed_obj[token_key]
                                        except json.JSONDecodeError:
                                            pass
                        elif isinstance(answer, list):
                            tokens = answer
                        elif isinstance(answer, dict):
                            # 如果 answer 已经是对象，直接提取
                            token_key = f"creator_tokens_{token_type}"
                            if token_key in answer:
                                tokens = answer[token_key]
                
                # 验证 tokens 是否为列表
                if isinstance(tokens, list):
                    logger.info(f"      ✓ 成功获取 {token_type} tokens: {len(tokens)} 个")
                    return tokens
                else:
                    error_detail = f"Dify API 返回的不是数组格式（{token_type} tokens），响应: {json.dumps(response_data, ensure_ascii=False)[:200]}"
                    logger.error(f"    {error_detail}")
                    raise DifyAPIError(error_detail)
                    
            except json.JSONDecodeError as e:
                error_detail = f"无法解析 Dify API 响应为 JSON（{token_type} tokens）: {str(e)}\n响应内容: {response.text[:500]}"
                logger.error(f"    {error_detail}")
                raise DifyAPIError(error_detail) from e
                
        except requests.exceptions.HTTPError as e:
            # HTTP 错误（4xx, 5xx）
            # 这种情况通常不会发生，因为我们在上面已经检查了 response.ok
            # 但如果 raise_for_status() 在其他地方被调用，这里会捕获
            error_detail = f"HTTP {e.response.status_code if e.response else 'Unknown'} {e.response.reason if e.response else ''}"
            if e.response:
                try:
                    error_body = e.response.json()
                    error_detail += f"\n响应内容: {json.dumps(error_body, ensure_ascii=False, indent=2)}"
                except (json.JSONDecodeError, ValueError):
                    error_detail += f"\n响应内容: {e.response.text[:500]}"
            logger.error(f"    调用 Dify API 失败（{token_type} tokens）: {error_detail}")
            logger.debug(f"    请求 URL: {api_url}")
            logger.debug(f"    请求 Payload: {json.dumps(payload, ensure_ascii=False, indent=2)}")
            raise DifyAPIError(f"Dify API 调用失败（{token_type} tokens）: {error_detail}") from e
        except requests.exceptions.RequestException as e:
            # 其他请求异常（网络错误、超时等）
            error_detail = f"{type(e).__name__}: {str(e)}"
            if hasattr(e, 'response') and e.response is not None:
                try:
                    error_body = e.response.json()
                    error_detail += f"\n响应内容: {json.dumps(error_body, ensure_ascii=False, indent=2)}"
                except (json.JSONDecodeError, ValueError):
                    error_detail += f"\n响应内容: {e.response.text[:500]}"
            logger.error(f"    调用 Dify API 失败（{token_type} tokens）: {error_detail}")
            logger.debug(f"    请求 URL: {api_url}")
            logger.debug(f"    请求 Payload: {json.dumps(payload, ensure_ascii=False, indent=2)}")
            raise DifyAPIError(f"Dify API 调用失败（{token_type} tokens）: {error_detail}") from e
        except DifyAPIError:
            # 重新抛出 DifyAPIError
            raise
        except Exception as e:
            # 其他异常
            import traceback
            error_detail = f"{type(e).__name__}: {str(e)}"
            logger.error(f"    处理 Dify API 响应失败（{token_type} tokens）: {error_detail}")
            logger.debug(f"    异常堆栈: {traceback.format_exc()}")
            logger.debug(f"    请求 URL: {api_url}")
            logger.debug(f"    请求 Payload: {json.dumps(payload, ensure_ascii=False, indent=2)}")
            raise DifyAPIError(f"处理 Dify API 响应失败（{token_type} tokens）: {error_detail}") from e
    
    def _generate_tokens_for_creator(self, creator_data: Dict[str, Any], creator_index: int = 0, total_count: int = 0) -> Tuple[Dict[str, Any], Optional[Exception]]:
        """
        为单个 Creator 生成 tokens（direct 和 implied）
        
        Args:
            creator_data: Creator 数据字典
            creator_index: Creator 索引（用于日志）
            total_count: 总数（用于日志）
            
        Returns:
            (creator_data_with_tokens, error): 包含 tokens 的 creator 数据和错误（如果有）
        """
        creator_name = creator_data.get("creator_name", "Unknown")
        creator_id = creator_data.get("creator_id", "Unknown")
        
        # 复制数据，避免修改原始数据
        result_data = creator_data.copy()
        
        # 如果 Dify 配置未设置，跳过 tokens 生成
        if not self.dify_url or not self.dify_api_key:
            logger.debug(f"    Creator {creator_index}/{total_count} ({creator_name}): Dify 配置未设置，跳过 tokens 生成")
            return result_data, None
        
        try:
            logger.info(f"    Creator {creator_index}/{total_count} ({creator_name}): 开始调用 Dify API 生成 tokens")
            
            # 并发调用 direct 和 implied（使用线程池）
            creator_tokens_direct = None
            creator_tokens_implied = None
            direct_error = None
            implied_error = None
            
            # 使用线程池并发执行 direct 和 implied
            with ThreadPoolExecutor(max_workers=2) as executor:
                future_direct = executor.submit(self._call_dify_for_tokens, creator_data, "direct")
                future_implied = executor.submit(self._call_dify_for_tokens, creator_data, "implied")
                
                # 获取结果
                try:
                    creator_tokens_direct = future_direct.result()
                except Exception as e:
                    direct_error = e
                    logger.warning(f"    Creator {creator_index}/{total_count} ({creator_name}): ✗ creator_tokens_direct 生成失败: {str(e)}")
                
                try:
                    creator_tokens_implied = future_implied.result()
                except Exception as e:
                    implied_error = e
                    logger.warning(f"    Creator {creator_index}/{total_count} ({creator_name}): ✗ creator_tokens_implied 生成失败: {str(e)}")
            
            # 设置结果
            if creator_tokens_direct:
                result_data["creator_tokens_direct"] = creator_tokens_direct
                logger.info(f"    Creator {creator_index}/{total_count} ({creator_name}): ✓ creator_tokens_direct 生成成功: {len(creator_tokens_direct)} 个")
            else:
                logger.warning(f"    Creator {creator_index}/{total_count} ({creator_name}): ⚠ creator_tokens_direct 为空")
            
            if creator_tokens_implied:
                result_data["creator_tokens_implied"] = creator_tokens_implied
                logger.info(f"    Creator {creator_index}/{total_count} ({creator_name}): ✓ creator_tokens_implied 生成成功: {len(creator_tokens_implied)} 个")
            else:
                logger.warning(f"    Creator {creator_index}/{total_count} ({creator_name}): ⚠ creator_tokens_implied 为空")
            
            # 如果有错误，返回第一个错误
            error = direct_error or implied_error
            if error:
                logger.warning(f"    Creator {creator_index}/{total_count} ({creator_name}): ⚠ 部分 tokens 生成失败")
            
            logger.info(f"    Creator {creator_index}/{total_count} ({creator_name}): ✓ Dify API 调用完成")
            return result_data, error
            
        except DifyAPIError as e:
            logger.warning(f"    Creator {creator_index}/{total_count} ({creator_name}): ✗ Dify API 调用失败: {str(e)}")
            return result_data, e
        except Exception as e:
            logger.error(f"    Creator {creator_index}/{total_count} ({creator_name}): ✗ 处理时发生异常: {str(e)}")
            return result_data, e
    
    def _generate_tokens_batch_concurrent(self, creators: List[Dict[str, Any]], max_workers: int = 10) -> List[Dict[str, Any]]:
        """
        并发生成 tokens（使用线程池，一次并发 max_workers 个请求）
        
        Args:
            creators: Creator 数据列表
            max_workers: 最大并发数（默认 10）
            
        Returns:
            包含 tokens 的 Creator 数据列表
        """
        if not creators:
            return []
        
        # 如果 Dify 配置未设置，直接返回原始数据
        if not self.dify_url or not self.dify_api_key:
            logger.info("Dify 配置未设置，跳过 tokens 生成")
            return creators
        
        logger.info(f"开始并发生成 tokens，并发数: {max_workers}，总数: {len(creators)}")
        
        processed_creators = []
        failed_count = 0
        
        # 使用线程池并发执行
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # 提交所有任务
            future_to_creator = {
                executor.submit(
                    self._generate_tokens_for_creator,
                    creator,
                    idx + 1,
                    len(creators)
                ): (idx, creator)
                for idx, creator in enumerate(creators)
            }
            
            # 收集结果（按原始顺序）
            results = [None] * len(creators)
            
            for future in as_completed(future_to_creator):
                idx, original_creator = future_to_creator[future]
                try:
                    creator_with_tokens, error = future.result()
                    results[idx] = creator_with_tokens
                    if error:
                        failed_count += 1
                except Exception as e:
                    logger.error(f"    处理 Creator 时发生未捕获的异常: {str(e)}")
                    results[idx] = original_creator
                    failed_count += 1
        
        # 过滤掉 None（理论上不应该有）
        processed_creators = [c for c in results if c is not None]
        
        logger.info(f"✓ Tokens 生成完成：成功 {len(processed_creators) - failed_count} 个，失败 {failed_count} 个")
        
        return processed_creators
    
    def _parse_sse_response(self, response: requests.Response) -> str:
        """
        解析 Dify 流式响应（Server-Sent Events 格式）
        返回完整的响应文本
        
        Args:
            response: requests Response 对象
            
        Returns:
            完整的响应文本
        """
        full_text = ""
        
        for line in response.iter_lines(decode_unicode=True):
            if not line:
                continue
            
            # SSE 格式：event: xxx 或 data: {...}
            if line.startswith("data: "):
                data_str = line[6:]  # 移除 "data: " 前缀
                
                # 跳过特殊事件
                if data_str == "[DONE]" or data_str.strip() == "":
                    continue
                
                try:
                    data = json.loads(data_str)
                    
                    # 根据 Dify API 的响应格式提取文本
                    if "answer" in data:
                        answer = data["answer"]
                        if isinstance(answer, str):
                            full_text += answer
                    elif "message" in data:
                        if isinstance(data["message"], dict) and "answer" in data["message"]:
                            full_text += data["message"]["answer"]
                        elif isinstance(data["message"], str):
                            full_text += data["message"]
                    elif "text" in data:
                        full_text += data["text"]
                    elif "content" in data:
                        full_text += data["content"]
                    elif isinstance(data, str):
                        # 如果 data 本身就是字符串（可能是 JSON 数组字符串）
                        full_text += data
                    elif isinstance(data, list):
                        # 如果 data 本身就是数组
                        return json.dumps(data, ensure_ascii=False)
                        
                except json.JSONDecodeError:
                    # 如果不是 JSON 格式，可能是纯文本数据
                    if data_str and not data_str.startswith("{"):
                        full_text += data_str
        
        return full_text.strip()
    
    def insert_to_supabase(self, creators: List[Dict[str, Any]], batch_size: int = 100) -> Dict[str, Any]:
        """
        批量插入 Creator 数据到 Supabase
        
        Args:
            creators: Creator 数据列表
            batch_size: 每批插入的数量
            
        Returns:
            插入结果字典，包含 total_inserted, total_failed, errors 等
            注意：即使部分插入失败，也不会抛出异常，而是返回错误信息
        """
        logger.info("开始批量插入数据到 Supabase...")
        logger.info(f"API URL: {self.api_url}")
        
        # 规范化所有 creator 的 content_category 字段，确保始终是数组格式
        for creator in creators:
            if "content_category" in creator:
                category = creator["content_category"]
                if isinstance(category, list):
                    # 确保列表中的元素都是字符串
                    creator["content_category"] = [str(c).strip() for c in category if str(c).strip()]
                else:
                    # 如果是字符串或其他类型，使用 _parse_categories 方法解析
                    parsed_categories = self._parse_categories(category)
                    creator["content_category"] = parsed_categories if parsed_categories else []
            else:
                # 如果没有 content_category 字段，设置为空列表
                creator["content_category"] = []
        
        total_inserted = 0
        total_failed = 0
        errors = []
        total_batches = (len(creators) + batch_size - 1) // batch_size
        
        logger.info(f"准备分 {total_batches} 批插入，每批最多 {batch_size} 条")
        
        for batch_idx, i in enumerate(range(0, len(creators), batch_size), start=1):
            batch = creators[i:i + batch_size]
            batch_range = f"{i+1}-{min(i+len(batch), len(creators))}"
            logger.info(f"处理第 {batch_idx}/{total_batches} 批 (行 {batch_range})，包含 {len(batch)} 条数据")
            
            try:
                logger.debug(f"  发送批量插入请求...")
                response = requests.post(
                    self.api_url,
                    headers=self.headers,
                    json=batch,
                    timeout=30
                )
                
                logger.debug(f"  响应状态码: {response.status_code}")
                
                if response.status_code in (200, 201):
                    response_data = response.json()
                    inserted_count = len(response_data) if response_data else len(batch)
                    total_inserted += inserted_count
                    logger.info(f"  ✓ 第 {batch_idx} 批插入成功，插入 {inserted_count} 条")
                else:
                    logger.warning(f"  ✗ 第 {batch_idx} 批批量插入失败 (状态码: {response.status_code})，尝试单个插入...")
                    logger.debug(f"  错误响应: {response.text[:500]}")
                    
                    # 尝试单个插入以获取更详细的错误信息
                    for creator_idx, creator in enumerate(batch, start=1):
                        creator_name = creator.get('creator_name', 'Unknown')
                        logger.debug(f"    尝试插入第 {creator_idx} 条: {creator_name}")
                        try:
                            single_response = requests.post(
                                self.api_url,
                                headers=self.headers,
                                json=creator,
                                timeout=30
                            )
                            if single_response.status_code in (200, 201):
                                total_inserted += 1
                                logger.debug(f"      ✓ {creator_name} 插入成功")
                            else:
                                # 记录失败但不终止，继续处理下一条
                                error_msg = single_response.text
                                total_failed += 1
                                error_detail = f"Creator '{creator_name}' - HTTP {single_response.status_code}: {error_msg}"
                                errors.append(error_detail)
                                logger.error(f"      ✗ {creator_name} 插入失败 (HTTP {single_response.status_code})")
                                logger.debug(f"        错误详情: {error_msg[:200]}")
                        except Exception as e:
                            # 记录异常但不终止，继续处理下一条
                            total_failed += 1
                            error_detail = f"Creator '{creator_name}' - 异常: {str(e)}"
                            errors.append(error_detail)
                            logger.error(f"      ✗ {creator_name} 插入异常: {str(e)}")
            except Exception as e:
                # 批量插入异常，记录错误但继续处理下一批
                total_failed += len(batch)
                error_detail = f"批次 {batch_idx} 批量插入异常: {str(e)}"
                errors.append(error_detail)
                logger.error(f"  ✗ 第 {batch_idx} 批批量插入异常: {str(e)}")
                logger.debug(f"    批次数据: {json.dumps(batch[:1], indent=2, ensure_ascii=False)}")  # 只打印第一条作为示例
        
        logger.info("=" * 80)
        logger.info("导入处理完成")
        logger.info(f"  总解析数据: {len(creators)} 条")
        logger.info(f"  成功插入: {total_inserted} 条")
        logger.info(f"  失败: {total_failed} 条")
        if errors:
            logger.warning(f"  错误数量: {len(errors)} 个")
            for error in errors[:5]:
                logger.warning(f"    - {error}")
        logger.info("=" * 80)
        
        return {
            "total_parsed": len(creators),
            "total_inserted": total_inserted,
            "total_failed": total_failed,
            "errors": errors[:10] if errors else []
        }
    
    def import_from_file(self, filename: str, contents: bytes) -> Dict[str, Any]:
        """
        从 Excel 文件导入 Creator 数据（完整流程）
        
        Args:
            filename: Excel 文件名
            contents: Excel 文件内容
            
        Returns:
            导入结果字典
            
        Raises:
            ExcelImportError: 如果导入过程中出现错误
        """
        logger.info("=" * 80)
        logger.info("开始处理 Excel 文件导入请求")
        logger.info(f"文件名: {filename}")
        
        # 验证文件
        self.validate_file(filename, contents)
        logger.info(f"✓ 文件格式验证通过")
        
        # 读取文件大小
        file_size = len(contents)
        logger.info(f"✓ 文件读取成功，大小: {file_size} 字节 ({file_size / 1024:.2f} KB)")
        
        # 解析 Excel
        headers, column_map, rows = self.parse_excel(contents)
        
        # 解析数据行
        creators = self.parse_data_rows(rows, column_map)
        
        # 插入到 Supabase
        result = self.insert_to_supabase(creators)
        
        return {
            "status": "success",
            "message": f"成功导入 {result['total_inserted']} 个 Creator，失败 {result['total_failed']} 个",
            **result
        }


def import_creators_from_excel(
    file_contents: bytes,
    filename: str,
    supabase_url: str,
    supabase_api_key: str,
    dify_url: str = None,
    dify_user: str = None,
    s3_bucket: str = None,
    s3_key_prefix: str = None
) -> Dict[str, Any]:
    """
    从 Excel 文件导入 Creator 数据的便捷函数
    
    Args:
        file_contents: Excel 文件内容（字节）
        filename: Excel 文件名
        supabase_url: Supabase URL
        supabase_api_key: Supabase API Key
        dify_url: Dify API URL (可选，从环境变量 DIFY_URL 获取)
        dify_user: Dify API User (可选，从环境变量 DIFY_USER 获取)
        s3_bucket: S3 Bucket 名称 (可选，从环境变量 S3_BUCKET 获取)
        s3_key_prefix: S3 键前缀 (可选，从环境变量 S3_CREATOR_PREFIX 获取，默认 "creator-signatures")
        
    Note:
        dify_api_key 将从环境变量 DIFY_API_KEY_TOKEN 自动获取
        
    Returns:
        导入结果字典
        
    Raises:
        ExcelImportError: 如果导入失败
    """
    # 从环境变量 DIFY_API_KEY_TOKEN 获取 dify_api_key
    dify_api_key = os.getenv("DIFY_API_KEY_TOKEN")
    
    importer = ExcelImporter(
        supabase_url=supabase_url,
        supabase_api_key=supabase_api_key,
        dify_url=dify_url,
        dify_api_key=dify_api_key,
        dify_user=dify_user,
        s3_bucket=s3_bucket,
        s3_key_prefix=s3_key_prefix
    )
    return importer.import_from_file(filename, file_contents)


def import_creators_from_json(
    json_file_path: str,
    supabase_url: str,
    supabase_api_key: str,
    dify_url: str = None,
    dify_user: str = None,
    s3_bucket: str = None,
    s3_key_prefix: str = None
) -> Dict[str, Any]:
    """
    从 output.json 文件导入 Creator 数据
    使用 JSON 文件中的 creator_id，导入成功的结果会写入缓存，避免多次执行时重复导入
    每个 Creator 在导入时会调用 Dify API 生成 creator_tokens_direct 和 creator_tokens_implied
    
    Args:
        json_file_path: output.json 文件路径
        supabase_url: Supabase URL
        supabase_api_key: Supabase API Key
        dify_url: Dify API URL (可选，从环境变量 DIFY_URL 获取)
        dify_user: Dify API User (可选，从环境变量 DIFY_USER 获取)
        s3_bucket: S3 Bucket 名称 (可选，从环境变量 S3_BUCKET 获取)
        s3_key_prefix: S3 键前缀 (可选，从环境变量 S3_CREATOR_PREFIX 获取)
        
    Note:
        dify_api_key 将从环境变量 DIFY_API_KEY_TOKEN 自动获取
        
    Returns:
        导入结果字典
        
    Raises:
        ExcelImportError: 如果导入失败
    """
    from pathlib import Path
    
    logger.info("=" * 80)
    logger.info("开始处理 JSON 文件导入请求")
    logger.info(f"文件路径: {json_file_path}")
    
    # 检查文件是否存在
    json_path = Path(json_file_path)
    if not json_path.exists():
        raise ExcelImportError(f"文件不存在: {json_file_path}")
    
    # 读取 JSON 文件
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            creators = json.load(f)
    except json.JSONDecodeError as e:
        raise ExcelImportError(f"JSON 文件格式错误: {str(e)}")
    except Exception as e:
        raise ExcelImportError(f"读取文件失败: {str(e)}")
    
    if not isinstance(creators, list):
        raise ExcelImportError("JSON 文件格式错误：根元素必须是数组")
    
    logger.info(f"✓ 文件读取成功，共 {len(creators)} 条记录")
    
    # 加载已导入的缓存
    cache_file = Path("cache") / "imported_creators.json"
    imported_creator_ids = set()
    
    if cache_file.exists():
        try:
            with open(cache_file, "r", encoding="utf-8") as f:
                cached_data = json.load(f)
                imported_creator_ids = set(cached_data.get("creator_ids", []))
                logger.info(f"✓ 加载缓存，已导入 {len(imported_creator_ids)} 个 Creator")
        except Exception as e:
            logger.warning(f"加载缓存失败: {str(e)}，将重新导入所有记录")
    
    # 过滤出未导入的记录（使用 JSON 文件中的 creator_id）
    new_creators = []
    skipped_count = 0
    
    for creator in creators:
        # 跳过空记录
        if not creator:
            skipped_count += 1
            continue
        
        creator_id = creator.get("creator_id")
        if not creator_id:
            logger.warning(f"跳过缺少 creator_id 的记录: {json.dumps(creator, ensure_ascii=False)[:100]}")
            skipped_count += 1
            continue
        
        creator_id_str = str(creator_id)
        if creator_id_str in imported_creator_ids:
            skipped_count += 1
            continue
        
        # 使用 JSON 文件中的 creator_id，不生成新的 UUID
        new_creators.append(creator)
    
    logger.info(f"✓ 过滤完成：新记录 {len(new_creators)} 条，已导入跳过 {skipped_count} 条")
    
    if not new_creators:
        logger.info("没有需要导入的新记录")
        return {
            "status": "success",
            "message": "没有需要导入的新记录",
            "total_parsed": len(creators),
            "total_inserted": 0,
            "total_skipped": skipped_count,
            "total_failed": 0,
            "errors": []
        }
    
    # 从环境变量获取 Dify API Key
    dify_api_key = os.getenv("DIFY_API_KEY_TOKEN")
    
    # 创建 ExcelImporter 实例（包含 Dify 配置，用于生成 tokens）
    importer = ExcelImporter(
        supabase_url=supabase_url,
        supabase_api_key=supabase_api_key,
        dify_url=dify_url,
        dify_api_key=dify_api_key,
        dify_user=dify_user,
        s3_bucket=s3_bucket,
        s3_key_prefix=s3_key_prefix
    )
    
    # 分批处理：每10条处理一次，立即写入数据库和缓存
    batch_size = 10
    total_batches = (len(new_creators) + batch_size - 1) // batch_size
    logger.info(f"开始分批处理，每批 {batch_size} 条，共 {total_batches} 批")
    logger.info("=" * 80)
    
    total_inserted = 0
    total_failed = 0
    all_errors = []
    
    for batch_idx in range(0, len(new_creators), batch_size):
        batch = new_creators[batch_idx:batch_idx + batch_size]
        batch_num = (batch_idx // batch_size) + 1
        logger.info(f"处理第 {batch_num}/{total_batches} 批 (第 {batch_idx+1}-{min(batch_idx+len(batch), len(new_creators))} 条)")
        
        try:
            # 1. 规范化 content_category 字段
            logger.debug(f"  规范化 content_category 字段...")
            for creator in batch:
                if "content_category" in creator:
                    category = creator["content_category"]
                    if isinstance(category, list):
                        creator["content_category"] = [str(c).strip() for c in category if str(c).strip()]
                    else:
                        parsed_categories = importer._parse_categories(category)
                        creator["content_category"] = parsed_categories if parsed_categories else []
                else:
                    creator["content_category"] = []
            
            # 2. 并发生成 tokens（一次并发10个请求）
            logger.info(f"  开始并发生成 tokens...")
            processed_batch = importer._generate_tokens_batch_concurrent(batch, max_workers=10)
            
            # 3. 插入到 Supabase
            logger.info(f"  开始插入到数据库...")
            batch_result = importer.insert_to_supabase(processed_batch, batch_size=len(processed_batch))
            
            # 4. 更新统计
            batch_inserted = batch_result.get("total_inserted", 0)
            batch_failed = batch_result.get("total_failed", 0)
            total_inserted += batch_inserted
            total_failed += batch_failed
            if batch_result.get("errors"):
                all_errors.extend(batch_result["errors"])
            
            # 5. 更新缓存：将成功导入的 creator_id 添加到缓存
            if batch_inserted > 0:
                inserted_creator_ids = set()
                for creator in processed_batch[:batch_inserted]:
                    creator_id = creator.get("creator_id")
                    if creator_id:
                        creator_id_str = str(creator_id)
                        inserted_creator_ids.add(creator_id_str)
                        logger.debug(f"    缓存 creator_id: {creator_id_str} (Creator: {creator.get('creator_name', 'Unknown')})")
                
                # 更新缓存集合
                imported_creator_ids.update(inserted_creator_ids)
                
                # 确保缓存目录存在
                cache_file.parent.mkdir(parents=True, exist_ok=True)
                
                # 保存缓存到文件
                try:
                    cache_data = {
                        "creator_ids": sorted(list(imported_creator_ids)),
                        "last_updated": datetime.now().isoformat(),
                        "total_count": len(imported_creator_ids),
                        "last_import": {
                            "inserted_count": len(inserted_creator_ids),
                            "timestamp": datetime.now().isoformat()
                        }
                    }
                    with open(cache_file, "w", encoding="utf-8") as f:
                        json.dump(cache_data, f, ensure_ascii=False, indent=2)
                    logger.info(f"  ✓ 第 {batch_num} 批处理完成：插入 {batch_inserted} 条，缓存已更新")
                except Exception as e:
                    logger.error(f"  ✗ 保存缓存失败: {str(e)}")
                    logger.exception(e)
            else:
                logger.info(f"  ✓ 第 {batch_num} 批处理完成：无新插入记录")
            
        except Exception as e:
            logger.error(f"  ✗ 第 {batch_num} 批处理失败: {str(e)}")
            logger.exception(e)
            total_failed += len(batch)
            all_errors.append(f"批次 {batch_num} 处理失败: {str(e)}")
            # 继续处理下一批，不中断整个流程
            continue
        
        logger.info("-" * 80)
    
    logger.info("=" * 80)
    logger.info("导入处理完成")
    logger.info(f"  总解析数据: {len(creators)} 条")
    logger.info(f"  新记录: {len(new_creators)} 条")
    logger.info(f"  成功插入: {total_inserted} 条")
    logger.info(f"  跳过（已导入）: {skipped_count} 条")
    logger.info(f"  失败: {total_failed} 条")
    if all_errors:
        logger.info(f"  错误数量: {len(all_errors)} 个")
    logger.info("=" * 80)
    
    return {
        "status": "success",
        "message": f"成功导入 {total_inserted} 个 Creator，跳过 {skipped_count} 个已导入的记录，失败 {total_failed} 个",
        "total_parsed": len(creators),
        "total_inserted": total_inserted,
        "total_skipped": skipped_count,
        "total_failed": total_failed,
        "errors": all_errors[:10] if all_errors else []
    }

